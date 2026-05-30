"""Parse substation bus terminal strings of the form  XXX<volts>kV[B<n>]_Term[(<dup>)]."""
from __future__ import annotations

import re
from dataclasses import dataclass

# (1) 3 alpha  (2) float volts + "kV"  (3) optional "B"+[1-9], then "_Term"  (4) optional "(n)"
_PATTERN = re.compile(r"([A-Za-z]{3})(\d+(?:\.\d+)?)kV(?:B([1-9]))?_Term(?:\((\d+)\))?")

# Voltage (compared as a number) -> the (i) digit used in the bus label. Default 0.
_I_MAP = {33: 3, 110: 7, 132: 8, 11: 1}


@dataclass(frozen=True)
class ParsedBus:
    source: str           # original string
    substation: str       # (1) the 3 alpha chars
    voltage_level: str    # (2) e.g. "33kV" / "5.5kV"
    name: str              # "BB" + i + j  (+ "_" + k if duplicate)


def parse_bus(s: str) -> ParsedBus | None:
    """Return a ParsedBus if `s` matches the format, else None."""
    m = _PATTERN.fullmatch(s)
    if m is None:
        return None

    substation, volts, n, k = m.groups()
    i = _I_MAP.get(float(volts), 0)               # (i): 3/7/8/1, else 0
    j = n if n is not None else "1"               # (j): n, or 1 if absent
    bus = f"BB{i}{j}" + (f"_{k}" if k is not None else "")  # (k) only if duplicate
    return ParsedBus(
        source=s,
        substation=substation,
        voltage_level=f"{volts}kV",
        name=bus,
    )


def parse_strings(strings):
    """Parse an iterable of strings.

    Returns (matched, unmatched): a list of ParsedBus and a list of the
    original strings that did not fit the format.
    """
    matched, unmatched = [], []
    for s in strings:
        parsed = parse_bus(s)
        if parsed is None:
            unmatched.append(s)
        else:
            matched.append(parsed)
    return matched, unmatched


if __name__ == "__main__":
    from openpyxl import load_workbook

    wb = load_workbook("/mnt/user-data/uploads/test_data.xlsx", read_only=True)
    ws = wb.active
    strings = [row[0] for row in ws.iter_rows(values_only=True) if row[0] is not None]

    matched, unmatched = parse_strings(strings)

    print(f"{len(strings)} strings | {len(matched)} matched | {len(unmatched)} unmatched\n")
    print(f"{'source':22} {'substation':11} {'voltage':9} {'bus'}")
    print("-" * 55)
    for p in matched:
        print(f"{p.source:22} {p.substation:11} {p.voltage_level:9} {p.bus}")

    print("\nSynthetic duplicate example:")
    print(" ", parse_bus("ABM33kVB1_Term(2)"))