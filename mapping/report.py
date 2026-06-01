"""
Excel reconciliation report writer.

Writes a ReconciliationResult to an .xlsx workbook with one worksheet per
status (matched / IPS-only / PF-only) plus a summary sheet, for offline
review and triage of the IPS -> PowerFactory mapping.
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for col_idx, header in enumerate(headers, start=1):
        width = len(str(header))
        for row in rows:
            val = row[col_idx - 1]
            if val is not None:
                width = max(width, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)


def write_reconciliation_report(result, out_dir, filename="reconciliation_report.xlsx"):
    """Write `result` to an .xlsx workbook in `out_dir`; return the file path."""
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, filename)

    wb = Workbook()

    # ---- Summary -----------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    for line in result.coverage_summary().splitlines():
        ws.append([line])
    ws.append([])
    ws.append(["Match tiers"])
    for tier, n in sorted(result.tier_counts().items()):
        ws.append([tier, n])
    ws.append([])
    ws.append(["IPS-only by element type"])
    for t, n in sorted(result.ips_only_by_element_type().items()):
        ws.append([t, n])
    ws.append([])
    ws.append(["PF-only by category"])
    for c, n in sorted(result.pf_only_by_category().items()):
        ws.append([c, n])
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12

    # ---- Matched -----------------------------------------------------------
    matched = sorted(result.matched, key=lambda m: (m.key.site_code, m.key.designation))
    _write_sheet(
        wb.create_sheet("Matched"),
        ["tier", "site", "voltage_kv", "designation", "n_settings",
         "setting_ids", "asset_names", "pf_raw_name", "pf_source"],
        [[m.tier, m.key.site_code, m.key.voltage_kv, m.key.designation,
          len(m.ips_devices), ";".join(m.setting_ids),
          ";".join(d.asset_name for d in m.ips_devices),
          m.pf.raw_name, m.pf.source] for m in matched],
    )

    # ---- IPS only ----------------------------------------------------------
    ips_only = sorted(result.ips_only.items(),
                      key=lambda kv: (kv[0].site_code, kv[0].designation))
    _write_sheet(
        wb.create_sheet("IPS only"),
        ["site", "voltage_kv", "designation", "element_type", "n_settings",
         "setting_ids", "asset_names", "location_path"],
        [[key.site_code, key.voltage_kv, key.designation,
          devs[0].element_type.value, len(devs),
          ";".join(d.setting_id for d in devs),
          ";".join(d.asset_name for d in devs),
          devs[0].location_path] for key, devs in ips_only],
    )

    # ---- PF only -----------------------------------------------------------
    pf_only = sorted(result.pf_only, key=lambda r: (r.key.site_code, r.key.designation))
    _write_sheet(
        wb.create_sheet("PF only"),
        ["site", "voltage_kv", "designation", "category", "raw_name", "source"],
        [[r.key.site_code, r.key.voltage_kv, r.key.designation,
          r.category, r.raw_name, r.source] for r in pf_only],
    )

    # ---- Ignored (substation not in PF) -----------------------------------
    ignored = sorted(result.ignored.items(),
                     key=lambda kv: (kv[0].site_code, kv[0].designation))
    _write_sheet(
        wb.create_sheet("Ignored (no PF site)"),
        ["site", "voltage_kv", "designation", "element_type", "n_settings",
         "setting_ids", "asset_names", "location_path"],
        [[key.site_code, key.voltage_kv, key.designation,
          devs[0].element_type.value, len(devs),
          ";".join(d.setting_id for d in devs),
          ";".join(d.asset_name for d in devs),
          devs[0].location_path] for key, devs in ignored],
    )

    wb.save(path)
    return path