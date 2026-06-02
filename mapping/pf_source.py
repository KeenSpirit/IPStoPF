"""
PowerFactory element source: produce canonical ``MappingKey`` references for
every relevant PowerFactory element, each carrying a pointer back to the
element and its relay cubicle for the eventual settings transfer.

Two sources:

* :func:`pf_refs_from_sites` - the production path. Consumes the
  ``domain.Site`` structure that ``process_pf_elements.process_elements``
  builds from the live PowerFactory model. Voltages and cubicles are taken
  directly from the model, so they are exact.

* :func:`pf_refs_from_workbook` - an offline validation path. Reads
  ``PowerFactory element data.xlsx`` and reconstructs keys using the element
  parsers, decoding winding/feeder voltages from terminal and designation
  names. Best effort: rows whose voltage cannot be decoded are reported as
  skipped rather than guessed.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional

from domain.mapping_key import MappingKey, VoltageKv
from process_pf_elements import pf_normalise as pn


@dataclass(frozen=True)
class PfElementRef:
    """A PowerFactory element resolved to a canonical key, with provenance."""
    key: MappingKey
    category: str                 # pn.CAT_* category
    raw_name: str                 # element loc_name (pre-normalisation)
    voltage_raw: VoltageKv        # voltage before int-normalisation
    source: str                   # grid / context, for traceability
    cubicle: object = None        # relay cubicle object (live); None offline


@dataclass
class PfSourceResult:
    refs: List[PfElementRef] = field(default_factory=list)
    skipped: List[str] = field(default_factory=list)   # rows that could not be keyed

    def by_key(self) -> Dict[MappingKey, List[PfElementRef]]:
        out: Dict[MappingKey, List[PfElementRef]] = {}
        for r in self.refs:
            out.setdefault(r.key, []).append(r)
        return out


# =============================================================================
# Production: from domain.Site objects
# =============================================================================

# domain ElementType.value -> pn category
_DOMAIN_TYPE_TO_CATEGORY = {
    "Busbar": pn.CAT_BUSBAR,
    "Switch": pn.CAT_SWITCH,
    "Capacitor bank": pn.CAT_CAP_BANK,
    "Feeder": pn.CAT_FEEDER,
    "Transformer": pn.CAT_TRANSFORMER,
    "Transformer HV": pn.CAT_TRANSFORMER,
    "Transformer LV": pn.CAT_TRANSFORMER,
    "Transformer LV A": pn.CAT_TRANSFORMER,
    "Transformer LV B": pn.CAT_TRANSFORMER,
    "Generator Cubicle": pn.GEN_CUBICLE
}

TRANSFORMER_HV = "Transformer HV"
TRANSFORMER_LV = "Transformer LV"
TRANSFORMER_LV_A = "Transformer LV A"
TRANSFORMER_LV_B = "Transformer LV B"
TRANSFORMER = "Transformer"
BUSBAR = "Busbar"
SWITCH = "Switch"
SPARE_SWITCH = "Spare switch"
CAPACITOR_BANK = "Capacitor bank"
FEEDER = "Feeder"
GEN_CUBICLE = "Generator Cubicle"


# When several elements at one (site, voltage) normalise to the same key, the
# reconciler takes pf_by_key[key][0]. Emit the transformer breaker bay
# (ElementType.TRANSFORMER) ahead of the winding bays so the breaker is matched.
_EMIT_PRIORITY = {"Transformer": 0}


def pf_refs_from_sites(sites) -> PfSourceResult:
    result = PfSourceResult()
    for site in sites:
        for voltage_kv, vl in site.voltage_levels.items():
            elems = [el for by_name in vl.elements.values() for el in by_name.values()]
            elems.sort(key=lambda el: _EMIT_PRIORITY.get(el.element_type.value, 99))
            for el in elems:
                category = _DOMAIN_TYPE_TO_CATEGORY.get(el.element_type.value)
                if category is None:
                    result.skipped.append(f"{site.name}:{el.name}:{el.element_type}")
                    continue
                key = pn.make_pf_key(site.name, category, el.name, voltage_kv)
                cub = getattr(getattr(el, "relay_cubicle", None), "obj", None)
                result.refs.append(PfElementRef(
                    key=key, category=category, raw_name=el.name,
                    voltage_raw=voltage_kv, source=site.name, cubicle=cub))
    return result


# =============================================================================
# Offline validation: from the spreadsheet
# =============================================================================

def _clean(x):
    return x.strip() if isinstance(x, str) else x


def _terminal_voltage(terminal_name: str):
    """Decode numeric kV from a bus terminal name via the bus parser."""
    from process_pf_elements import bus_parser as bp
    if not terminal_name:
        return None
    p = bp.parse_bus(terminal_name.strip())
    if p is None:
        return None
    try:
        return pn.normalise_voltage_kv(float(p.voltage_level.replace("kV", "")))
    except ValueError:
        return None


def _terminal_site(terminal_name: str):
    from process_pf_elements import bus_parser as bp
    if not terminal_name:
        return None
    p = bp.parse_bus(terminal_name.strip())
    return p.substation if p else None


def pf_refs_from_workbook(path: str) -> PfSourceResult:
    """Build PF references from PowerFactory element data.xlsx (offline)."""
    from openpyxl import load_workbook
    from process_pf_elements import (
        bus_parser as bp, cap_bank_parser as cbp,
        switch_parser as swp, tfmr_parser as tp, line_parser as lp,
    )

    wb = load_workbook(path, read_only=True, data_only=True)
    result = PfSourceResult()

    def rows(sheet):
        return list(wb[sheet].iter_rows(values_only=True))[1:]

    def add(site, category, raw_name, voltage, source):
        if site is None or voltage is None or raw_name is None:
            result.skipped.append(f"{source}:{raw_name}")
            return
        key = pn.make_pf_key(site, category, str(raw_name), voltage)
        result.refs.append(PfElementRef(key=key, category=category,
                                        raw_name=str(raw_name), voltage_raw=voltage,
                                        source=str(source)))

    # Busbars
    for name, volt, grid in rows("busbar"):
        name = _clean(name)
        p = bp.parse_bus(name) if name else None
        if p is None:
            result.skipped.append(f"busbar:{name}")
            continue
        add(p.substation, pn.CAT_BUSBAR, p.name, volt, grid)

    # Cap banks
    for name, term, volt, grid in rows("cap bank"):
        name = _clean(name)
        p = cbp.parse_cap_bank(name) if name else None
        if p is None:
            result.skipped.append(f"cap bank:{name}")
            continue
        add(p.substation, pn.CAT_CAP_BANK, p.name, volt, grid)

    # Switches (ElmCoup): bus couplers and line switches
    for name, ti, tj, grid in rows("switch"):
        name = _clean(name)
        p = swp.parse_switch(name) if name else None
        if p is None:
            result.skipped.append(f"switch:{name}")
            continue
        add(p.substation, pn.CAT_SWITCH, p.name, p.voltage_level, grid)

    # Lines (feeders): site from a terminal, voltage from the F-number
    for elmlne, ti, tj, grid in rows("line"):
        s = _clean(elmlne)
        if s is None:
            continue
        fname = lp.extract_leading_number(str(s)) or str(s)
        site = _terminal_site(ti) or _terminal_site(tj)
        volt = pn.feeder_voltage_from_name(fname)
        if volt is None:
            volt = _terminal_voltage(ti) or _terminal_voltage(tj)
        add(site, pn.CAT_FEEDER, fname, volt, grid)

    # Two-winding transformers
    for name, hv, lv, grid in rows("tr2"):
        name = _clean(name)
        p = tp.parse_tfmr(name) if name else None
        if p is None:
            result.skipped.append(f"tr2:{name}")
            continue
        for term in (hv, lv):
            add(p.substation, pn.CAT_TRANSFORMER, p.name, _terminal_voltage(term), grid)

    # Three-winding transformers
    for name, hv, mv, lv, grid in rows("tr3"):
        name = _clean(name)
        p = tp.parse_tfmr(name) if name else None
        if p is None:
            result.skipped.append(f"tr3:{name}")
            continue
        for term in (hv, mv, lv):
            add(p.substation, pn.CAT_TRANSFORMER, p.name, _terminal_voltage(term), grid)

    return result